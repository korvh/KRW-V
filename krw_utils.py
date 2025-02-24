from causallearn.search.FCMBased.lingam import DirectLiNGAM
from causallearn.search.ConstraintBased.PC import pc
from causallearn.search.ConstraintBased.FCI import fci
from causallearn.utils.PCUtils.BackgroundKnowledge import BackgroundKnowledge
from causallearn.graph.GeneralGraph import GeneralGraph
from causallearn.graph.GraphNode import GraphNode
from causallearn.utils.GraphUtils import GraphUtils
import networkx as nx
import pydot
import io
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import matplotlib.patches as mpatches
from matplotlib.colors import LinearSegmentedColormap
import seaborn as sns
import hiplot as hip
import pandas as pd
from sklearn.base import TransformerMixin, BaseEstimator
from sklearn.base import BaseEstimator, RegressorMixin
from sklearn.utils import check_array
from scipy.stats import yeojohnson, boxcox
from openpyxl import load_workbook
from sklearn.preprocessing import OneHotEncoder
from sklearn.metrics import mean_squared_error, mean_absolute_error, make_scorer, r2_score


# Function to load all data into a single dataframe
def read_excel_with_multiple_sheets(path_name, file_name, sheets_to_skip=[]):
    raw_df = pd.DataFrame()
    full_path = path_name + '/' + file_name
    wb = load_workbook(filename=full_path)
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name not in sheets_to_skip:
            sheet_df = pd.read_excel(full_path, sheet_name=sheet_name)
            raw_df = pd.concat([raw_df, sheet_df], axis='index')
    return raw_df


# Helper functions for building regression model pipeline
# Custom function for log-transform (and inverse)
def log_transform(x):
    return np.log1p(x)

def inv_log_transform(x):
    return np.expm1(x)

# Custom function for signed log1p-transform (and inverse)
def signed_log1p_transform(x):
    return np.sign(x) * np.log1p(np.abs(x))

def inv_signed_log1p_transform(x):
    return np.sign(x) * np.expm1(np.abs(x))

# Custom function to convert numeric values to strings
def safe_convert_to_str(X):
    return X.applymap(lambda x: str(x) if x is not np.nan else x)

# Custom function to drop columns by name
def drop_columns_by_name(X, columns_to_drop, drop=True):
    if drop:
        X = X.drop(columns=columns_to_drop, errors='ignore')
    return X

# Custom function to drop rows with ones in y
def drop_rows_with_ones(X, y=None):
    if y is not None:
        mask = y!=1
        return X[mask], y[mask]
    return X

# Custom function to drop rows with nans in either X or y
def drop_rows_with_nans(X, y=None):
    nan_mask_X = np.isnan(X).any(axis=1)
    if y is not None:
        nan_mask_y = np.isnan(y)
        nan_mask = nan_mask_X | nan_mask_y
        X_clean = X[~nan_mask]
        y_clean = y[~nan_mask]
        return X_clean, y_clean
    else:
        X_clean = X[~nan_mask_X]
        return X_clean

# Function to create monotonic constraints for specific features
def monotonic_constraints(num_vars, ord_vars, cat_vars, cat_encode, X, features_increasing, features_decreasing):
    new_column_order = num_vars
    new_column_order.extend(ord_vars)
    if cat_encode != 'OHE':
        new_column_order.extend(cat_vars)
    else:
        for col in cat_vars:
            ohe = OneHotEncoder(sparse=False)
            ohe.fit_transform(X[[col]])
            new_column_order.extend(ohe.categories_[0])
    monotone_constraints = []
    for col in new_column_order:
        if col in features_increasing:
            monotone_constraints.append(1)
        elif col in features_decreasing:
            monotone_constraints.append(-1)
        else:
            monotone_constraints.append(0)
    return monotone_constraints, new_column_order


# Classes for custom (target) transformers
class YeoJohnsonTargetTransformer(BaseEstimator, TransformerMixin):
    """
    A transformer for applying the Yeo-Johnson transformation to a target variable with optional predefined lambda values.

    Parameters
    ----------
    lmbda : float, optional
        If provided, this lambda value will be used for the transformation. If None, the lambda value is estimated from the data.
    replace_nan_percentile : float, optional (default=99)
        The percentile value used to replace NaN values in the inverse transformation. Should be between 0 and 100.
    """

    def __init__(self, lmbda: float = None, replace_nan_percentile: float = 99):
        """
        Initialize the transformer with optional lambda and replace_nan_percentile.

        Parameters
        ----------
        lmbda : float, optional
            If provided, this lambda value will be used for the transformation. If None, the lambda value is estimated from the data.
        replace_nan_percentile : float, optional (default=99)
            The percentile value used to replace NaN values in the inverse transformation. Should be between 0 and 100.
        """
        self.lmbda = lmbda
        self.lmbda_ = None
        self.replace_nan_percentile = replace_nan_percentile

    def fit(self, y, *_):
        """
        Fit the transformer by estimating the lambda parameter if not provided.

        Parameters
        ----------
        y : array-like, shape (n_samples,)
            The data to fit.
        
        Returns
        -------
        self : object
            Returns self.
        """
        y = check_array(y, ensure_2d=False)
        if self.lmbda is None:
            _, self.lmbda_ = yeojohnson(y)
        else:
            self.lmbda_ = self.lmbda
        return self

    def transform(self, y, *_):
        """
        Apply the Yeo-Johnson transformation to the data.

        Parameters
        ----------
        y : array-like, shape (n_samples,)
            The data to transform.

        Returns
        -------
        y_trans : array, shape (n_samples,)
            The transformed data.
        """
        y = check_array(y, ensure_2d=False)
        return yeojohnson(y, lmbda=self.lmbda_).flatten()

    def inverse_transform(self, y, *_):
        """
        Apply the inverse of the Yeo-Johnson transformation to the data.

        Parameters
        ----------
        y : array-like, shape (n_samples,)
            The data to inverse transform.

        Returns
        -------
        y_inv : array, shape (n_samples,)
            The inverse transformed data, with NaN values replaced by the specified percentile of the non-NaN results across a range of values.
        """
        y = check_array(y, ensure_2d=False)
        y_inv = self._inverse_yeojohnson(y, self.lmbda_)

        # Identify NaN values in the transformed output
        nan_mask = np.isnan(y_inv)
        if np.any(nan_mask):
            # Create a range of values based on min and max of y
            y_range = np.linspace(y.min(), y.max(), 1000)
            y_range_inv = self._inverse_yeojohnson(y_range, self.lmbda_)

            # Sort the inverse transformed range values
            y_range_inv = np.sort(y_range_inv)

            # Compute the specified percentile of the non-NaN values
            non_nan_values = y_range_inv[~np.isnan(y_range_inv)]
            if non_nan_values.size > 0:
                percentile_value = np.percentile(non_nan_values, self.replace_nan_percentile)
                y_inv[nan_mask] = percentile_value

        return y_inv.flatten()

    def fit_transform(self, y, *_):
        """
        Fit the transformer and apply the Yeo-Johnson transformation to the data.

        Parameters
        ----------
        y : array-like, shape (n_samples,)
            The data to fit and transform.

        Returns
        -------
        y_trans : array, shape (n_samples,)
            The transformed data.
        """
        return self.fit(y).transform(y)

    def _inverse_yeojohnson(self, y, lmbda):
        """
        Apply the inverse of the Yeo-Johnson transformation to the data.

        Parameters
        ----------
        y : array-like, shape (n_samples,)
            The data to inverse transform.

        lmbda : float
            The lambda value used in the Yeo-Johnson transformation.

        Returns
        -------
        out : array, shape (n_samples,)
            The inverse transformed data.
        """
        pos = y >= 0
        out = np.zeros_like(y)
        if lmbda == 0:
            out[pos] = np.expm1(y[pos])
            out[~pos] = -np.expm1(-y[~pos])
        else:
            out[pos] = np.power(y[pos] * lmbda + 1, 1 / lmbda) - 1
            out[~pos] = 1 - np.power(-(2 - lmbda) * y[~pos] + 1, 1 / (2 - lmbda))
        return out

    def get_params(self, deep=True):
        """
        Get parameters for this estimator.

        Parameters
        ----------
        deep : bool, optional
            If True, will return the parameters for this estimator and contained subobjects that are estimators.

        Returns
        -------
        params : dict
            Parameter names mapped to their values.
        """
        return {"lmbda": self.lmbda_, "replace_nan_percentile": self.replace_nan_percentile}


class LogTargetTransformer(BaseEstimator, TransformerMixin):
    """
    Custom transformer for applying log1p transformation to target variables.

    This transformer applies a log1p transformation (log(1 + x)) to the target
    variable to handle zero and small positive values gracefully. It also
    implements the inverse transformation using expm1 (exp(x) - 1).

    Attributes:
        min_ (float): Minimum value of the target variable, used for shifting
                      values before applying the log1p transformation.
    """

    def __init__(self):
        """
        Initializes the LogTransformer.
        """
        self.min_ = None

    def fit(self, y):
        """
        Fit the transformer to the target variable.

        Parameters
        ----------
        y : array-like, optional
            Target variable.

        Returns
        -------
        self : object
            Returns an instance of self.
        """
        y = self._convert_to_numpy(y)
        self.min_ = y.min() if np.any(y <= 0) else 0
        return self

    def transform(self, y):
        """
        Apply the log1p transformation to the target variable.

        Parameters
        ----------
        y : array-like, optional
            Target variable.

        Returns
        -------
        y_trans : array-like
            Transformed target variable.
        """
        y = self._convert_to_numpy(y)
        return np.log1p(y - self.min_)

    def fit_transform(self, y):
        """
        Fit to data, then transform it.

        Parameters
        ----------
        y : array-like, optional
            Target variable.

        Returns
        -------
        y_trans : array-like
            Transformed target variable.
        """
        y = self._convert_to_numpy(y)
        return self.fit(y).transform(y)

    def inverse_transform(self, y):
        """
        Apply the inverse transformation (expm1) to the transformed target variable.

        Parameters
        ----------
        y : array-like, optional
            Transformed target variable.

        Returns
        -------
        y_inv : array-like
            Original scale target variable.
        """
        return np.expm1(y) + self.min_

    def _convert_to_numpy(self, y):
        """
        Convert input to numpy array if it's a pandas Series.

        Parameters
        ----------
        y : array-like or pandas Series

        Returns
        -------
        y_numpy : numpy array
        """
        if isinstance(y, pd.Series):
            return y.values
        return check_array(y, ensure_2d=False)


# Custom transformer class for log1p-transform with min shift
class Log1pTransformerWithShift(BaseEstimator, TransformerMixin):
    def __init__(self, shift=None):
        self.shift_ = shift
    
    def fit(self, X, y=None):
        # Calculate shift per feature if not provided during initialization
        if self.shift_ is None:
            self.shift_ = np.min(X, axis=0) - 1 if np.any(np.min(X, axis=0) < 0) else np.zeros(X.shape[1])
        return self
    
    def transform(self, X):
        return np.log1p(X - self.shift_)
    
    def inverse_transform(self, X):
        return np.expm1(X) + self.shift_

    def get_params(self, deep=True):
        return {"shift": self.shift_}

    # def set_params(self, **params):
    #     for key, value in params.items():
    #         setattr(self, key, value)
    #     return self


# Get regression metrics
def get_regression_metrics(forecast, actual, baseline=None, Q_nrmse=0, b_range=0.1, is_baseline=False, bounds=(0.,1.)):
    # Mean Error (ME)
    # Mean Absolute Error (MAE)
    # Root Mean Squared Error (RMSE)
    # Root Mean Squared Log Error (RMSLE)
    # Correlation between the y and the ypred (CORR)
    # Mean absolute scaled error (MASE)
    # Root Relative Squared Error (RRSE)
    # Coefficient of Determination (CoD)
    # R-squared (R2)
    # Concordance-Correlation_Coefficient (CCC)
    
    EPSILON = 1e-10
    digits = 4
    
    y = actual.copy()
    if isinstance(y, pd.Series):
        y = y.to_numpy().flatten()
    ypred = forecast.copy()
    if isinstance(ypred, pd.Series):
        ypred = ypred.to_numpy().flatten()
    if baseline is not None:
        ybase = baseline.copy()
        if isinstance(ybase, pd.Series):
            ybase = ybase.to_numpy().flatten()
    if bounds is not None:
        mask = np.where((y>=bounds[0]) & (y <=bounds[1]))
        y, ypred = y[mask], ypred[mask]
        if baseline is not None:
            ybase = ybase[mask]

    me_ = np.round(np.nanmean(y - ypred), digits)
    mae_ = np.round(np.nanmean(np.abs(y - ypred)), digits)
    rmse_ = np.round(np.nanmean((y - ypred) ** 2) ** 0.5, digits)
    # # RMLSE penalizes overestimates and underestimates equally, but meassured in a different way. It gives equal penalties to equal ratios, regardless 
    # # whether it is an overestimate or an underestimate (https://medium.com/analytics-vidhya/root-mean-square-log-error-rmse-vs-rmlse-935c6cc1802a)
    # if (np.min(y) >= 0) & (np.min(ypred) >= 0):
    #     rmsle_ = np.round(np.sqrt(mean_squared_log_error(np.nan_to_num(y, copy=True, nan=0.0, posinf=None, neginf=None), np.nan_to_num(ypred, copy=True, nan=0.0, posinf=None, neginf=None))), digits)
    # else:
    #     rmsle_ = np.nan
    if Q_nrmse == 0:
        if (np.nanmax(y) - np.nanmin(y) == 0) or (np.isnan(np.nanmax(y) - np.nanmin(y))):
            nrmse = np.nan
        else:
            nrmse = np.round((np.nanmean((y - ypred) ** 2) ** 0.5) / (np.nanmax(y) - np.nanmin(y)), digits)
    else:
        if (np.nanquantile(y, 1-Q_nrmse) - np.nanquantile(y, Q_nrmse) == 0) or (np.isnan(np.nanquantile(y, 1-Q_nrmse) - np.nanquantile(y, Q_nrmse))):
            nrmse = np.nan
        else:
            nrmse = np.round((np.nanmean((y - ypred) ** 2) ** 0.5) / (np.nanquantile(y, 1-Q_nrmse) - np.nanquantile(y, Q_nrmse)), digits)
    r2_ = np.round(r2_score(np.nan_to_num(y, copy=True, nan=0.0, posinf=None, neginf=None), np.nan_to_num(ypred, copy=True, nan=0.0, posinf=None, neginf=None)), digits)
    if (np.nanvar(y) == 0) or (np.isnan(np.nanvar(y))):
        cod_ = np.nan
    else:
        cod_ = np.round(1 - ((np.nanmean((y - ypred) ** 2) / np.nanvar(y)) * (len(y) / (len(y) - 1))), digits)
    corr_ = np.round(pd.DataFrame({'y':y, 'ypred':ypred}).corr().values[0][1], digits)
    if baseline is None:
        mase_ = np.nan
    elif is_baseline:
        mase_ = np.nan
    else:
        mase_ = np.round(np.nanmean(np.abs(y - ypred)) / np.nanmean(np.abs(y - ybase)), digits)

    if baseline is None:
        rrse_ = np.nan
    elif is_baseline:
        rrse_ = np.nan
    else:
        rrse_ = np.round(np.sqrt(np.nansum((y - ypred) ** 2)) / np.sqrt(np.nansum((y - ybase) ** 2)), digits)
    
    ccc = concordance_correlation_coefficient(y, ypred)

    df_diff = pd.DataFrame()
    df_diff['dist'] = np.abs(y - ypred)
    df_diff['within'] = df_diff['dist'] <= b_range
    in_range = np.round(df_diff['within'].sum() / df_diff.shape[0], digits)

    metrics = {
        'RMSE':rmse_, 
        'RRSE':rrse_, 
        # 'RMSLE':rmsle_,
        'MAE':mae_, 
        'MASE':mase_, 
        'ME':me_,
        'R2':r2_, 
        'CoD':cod_, 
        '+/-0.1':in_range,
        'nRMSE':nrmse, 
        'Corr':corr_, 
        'CCC':ccc,
    }
    return metrics


# Concordance Correlation Coefficient
def concordance_correlation_coefficient(y_true, y_pred, sample_weight=None, multioutput='uniform_average'):
    """Concordance correlation coefficient.
    
    See: https://rowannicholls.github.io/python/statistics/agreement/concordance_correlation_coefficient.html
         https://github.com/stylianos-kampakis/supervisedPCA-Python/blob/master/Untitled.py

    The concordance correlation coefficient is a measure of inter-rater agreement.
    It measures the deviation of the relationship between predicted and true values
    from the 45 degree angle.

    Read more: https://en.wikipedia.org/wiki/Concordance_correlation_coefficient
    Original paper: Lawrence, I., and Kuei Lin. "A concordance correlation coefficient to evaluate reproducibility." Biometrics (1989): 255-268.  

    Parameters
    ----------
    y_true : array-like of shape = (n_samples) or (n_samples, n_outputs)
        Ground truth (correct) target values.

    y_pred : array-like of shape = (n_samples) or (n_samples, n_outputs)
        Estimated target values.

    Returns
    -------
    loss : A float in the range [-1,1]. A value of 1 indicates perfect agreement
    between the true and the predicted values.
    """
    cor = np.corrcoef(y_true, y_pred)[0][1]
    mean_true = np.mean(y_true)
    mean_pred = np.mean(y_pred)
    var_true = np.var(y_true)
    var_pred = np.var(y_pred)
    sd_true = np.std(y_true)
    sd_pred = np.std(y_pred)
    numerator = 2 * cor * sd_true * sd_pred
    denominator = var_true + var_pred + (mean_true - mean_pred)**2
    return (numerator / denominator)


# Custom function for permutation importance (when using models from sklearn_quantile)
def custom_permutation_importance(model, X, y, metric=mean_absolute_error, n_repeats=5, random_state=42):
    baseline_score = metric(y, model.predict(X.values))
    importances = np.zeros(X.shape[1])

    for i, col in enumerate(X.columns):
        score_diffs = []
        for _ in range(n_repeats):
            X_permuted = X.copy()
            X_permuted[col] = np.random.permutation(X_permuted[col])
            permuted_score = metric(y, model.predict(X_permuted.values))
            score_diffs.append(permuted_score - baseline_score)
        importances[i] = np.mean(score_diffs)

    return importances


# Function to plot model performance
def plot_model_performance(ytrue_train, ypred_train, ytrue_test, ypred_test, width, height, ncols=3, dpi=150, binsize=0.2, error_range=0.1, fontsize=7):
    fig, axes = plt.subplots(figsize=(width, height), nrows=1, ncols=ncols, dpi=dpi)
    fig.subplots_adjust(wspace=0.4)
    k = 0
    ax = axes.flatten()[k]
    ax.scatter(ytrue_train, ypred_train, s=4, label='Training', zorder=10)
    ax.scatter(ytrue_test, ypred_test, s=4, label='Test', zorder=20)
    ax.set_xlabel('EKR score')
    ax.set_ylabel('Voorspelde score')
    x = np.linspace(-0.05, 1.05, 100)
    y = x
    ax.plot(x, y, '--', lw=0.3, c='dimgray', zorder=0)
    ax.fill_between(x, y - error_range, y + error_range, color='lightgray', alpha=0.5, zorder=-10)
    ax.set_xlim(-0.05, 1.05)
    ax.set_ylim(-0.05, 1.05)
    ax.legend(loc='upper left', fontsize=fontsize, frameon=False, handletextpad=0.3)
    ax.xaxis.set_tick_params(labelsize=fontsize)
    ax.yaxis.set_tick_params(labelsize=fontsize)
    ax.set_xlabel(ax.get_xlabel(), fontsize=fontsize)
    ax.set_ylabel(ax.get_ylabel(), fontsize=fontsize)
    k += 1
    ax = axes.flatten()[k]
    result_df = pd.DataFrame(
        {
            'ytrue':ytrue_test,
            'error':ypred_test - ytrue_test,
        }
    )
    bins = np.arange(0, 1 + binsize, binsize)
    labels = [f"{round(bins[i], 1)} - {round(bins[i+1], 1)}" for i in range(len(bins)-1)]  # Labels for bins
    result_df['binned'] = pd.cut(result_df['ytrue'], bins=bins, labels=labels, include_lowest=True)
    result_df = result_df.assign(
        **{
            f"< -{error_range}": (result_df['error'] < -error_range) * 1,
            f"+/- {error_range}": (result_df['error'].between(-error_range, error_range)) * 1,
            f"> {error_range}": (result_df['error'] > error_range) * 1,
        }
    )
    result_df = result_df.drop(columns=['ytrue', 'error'])
    result_df = result_df.groupby('binned', observed=False).sum()
    colors = ['firebrick', '#2ca02c', 'darkorange'] # colors for stacked bars, using matplotlib default green
    result_df.plot(kind='bar', stacked=True, color=colors, ax=ax)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles[::-1], labels[::-1], fontsize=fontsize, frameon=False, handlelength=.8, handletextpad=0.5, title='Voorspelfout', title_fontsize=fontsize)
    ax.set_xlabel('EKR score', fontsize=fontsize)
    ax.set_ylabel('Aantal datapunten', fontsize=fontsize)
    ax.set_xticks(ax.get_xticks())
    ax.set_xticklabels(ax.get_xticklabels(), rotation=30, ha='right')
    ax.xaxis.set_tick_params(labelsize=fontsize)
    ax.yaxis.set_tick_params(labelsize=fontsize)
    for ax in axes.flatten()[k+1:]:
        ax.axis('off')
    return fig


# Function to plot SHAP dependence plots
def shap_dependence_plots_with_target_as_legend(shap_data, X, y, shap_features, num_cols=4, share_yaxes=True, cmap='RdYlBu', markersize=25, dpi=150):
    # Create figure and axes
    num_rows = int(np.ceil(shap_data.shape[1] / num_cols))
    max_cols = min(num_cols, shap_data.shape[1])
    fig, axes = plt.subplots(figsize=(16,3.5*np.ceil(shap_data.shape[1]) / num_cols), ncols=num_cols, nrows=num_rows, sharey=share_yaxes, dpi=dpi)
    fig.subplots_adjust(bottom=0.05, top=0.95, left=0.05, right=0.9, wspace=0.5, hspace=0.6)
    # Get data to plot
    shap_to_plot = pd.DataFrame(shap_data, columns=shap_features)
    miny, maxy = np.inf, -np.inf
    for k, feat_name in enumerate(shap_features):
        df_plot = X.copy()[[feat_name]]
        df_plot = df_plot.assign(
            **{
                'SHAP value':shap_to_plot[feat_name].values,
                'label':y.values
            }
        )
        ax = axes.flatten()[k]
        show_legend = True if k==max_cols-1 else False
        sns.scatterplot(data=df_plot, x=feat_name, y='SHAP value', hue='label', alpha=1, palette=cmap, legend=show_legend, s=markersize, ax=ax)
        ax.set_xlabel(feat_name, fontsize=12)
        ax.set_ylabel(f"SHAP value for\n'{feat_name}'", fontsize=12)
        ax.tick_params(axis='both', labelsize=12)
        miny, maxy = min(miny, ax.get_ylim()[0]), max(maxy, ax.get_ylim()[1])
        # Change axes properties
        ax.spines['left'].set_color('dimgray')
        ax.spines['bottom'].set_color('dimgray')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_facecolor('white')
        ax.grid(axis='both', color=[.7,.7,.7], linestyle='-', linewidth=.5)
        ax.set_axisbelow(True)
        ax.set_ylabel('SHAP waarde')
        plt.draw()
        if df_plot[feat_name].dtype in ['object','category']:
            ax.set_xticks(ax.get_xticks())
            ax.set_xticklabels(ax.get_xticklabels(), rotation=30, ha='right')
    # Set limits of shared yaxes
    if share_yaxes:
        for ax in axes.flatten():
            ax.set_ylim(miny,maxy)
    # Show legend next to top right axis
    legend = axes.flatten()[max_cols-1].legend(bbox_to_anchor=(1,1.05), loc='upper left', frameon=True, fontsize=12, title='EKR score')
    legend.get_title().set_fontsize('14')
    legend.get_frame().set_facecolor('white')
    # Delete unused axes
    if k < len(axes.flatten()):
        if num_rows > 1:
            for ax_del in axes.flatten()[k+1:]:
                fig.delaxes(ax=ax_del)
        else:
            for ax in axes.flatten()[k+1:]:
                ax.axis('off')
    return fig
   

# Function to plot a causal graph
def plot_causal_graph(G, node_names, ax, ci_type, color_direction=False, show_edge_width=False, max_line_width=5, arrow_dir='forward', graph_type='causal', adj_matrix=[], title=None, fontsize=24, pyd_layout='dot', bbox_to_anchor=(1,1), loc_legend='upper left', dpi=100):
    # Check if G has the attributes typical of a causal-learn graph
    is_causal_learn_graph = True
    try:
        pyd = GraphUtils.to_pydot(G, labels=node_names)
    except:
        is_causal_learn_graph = False

    if is_causal_learn_graph:
        # G is a causal-learn graph
        pyd = GraphUtils.to_pydot(G, labels=node_names)
        pyd.set_layout(pyd_layout)
    else:
        # if not a causal-learn graph, assume it's a networkx DiGraph
        pyd = pydot.Dot(graph_type='digraph', layout=pyd_layout)
        if (len(adj_matrix) > 0) & (len(adj_matrix[np.abs(adj_matrix) > 0]) > 0):
            max_strength = np.nanmax(np.abs(adj_matrix))
            min_strength = np.nanmin(np.abs(adj_matrix[np.abs(adj_matrix) > 0]))
        for i, edge in enumerate(G.edges()):
            # Get strength of causal link from adjacency matrix
            cause = edge[0]
            effect = edge[1]
            cause_idx = node_names.index(cause)
            effect_idx = node_names.index(effect)
            if len(adj_matrix) > 0:
                link_strength = adj_matrix[cause_idx, effect_idx]
            # Set color for direction of causal link
            if color_direction:
                if link_strength > 0:
                    color='blue'
                else:
                    color='firebrick'
            else:
                color='black'
            # Set width for strength of causal link
            if show_edge_width:
                width = 0.3 + (max_line_width* abs(link_strength) / (max_strength - min_strength))
            else:
                width = 0.8
            # Set edge style
            if arrow_dir == 'both':
                edge_style = pydot.Edge(cause, effect, color=color, penwidth=width, dir='both')
            elif arrow_dir == 'none':
                edge_style = pydot.Edge(cause, effect, color=color, penwidth=width, arrowhead='none')
            else:
                edge_style = pydot.Edge(cause, effect, color=color, penwidth=width)  # Default: forward arrow
            # Add edge to pydot graph
            pyd.add_edge(edge_style)

        # Customize node shapes (ellipses for causal-learn style)
        for node_name in node_names:
            pyd.add_node(pydot.Node(node_name, shape="ellipse"))

    # Convert pydot graph to png
    tmp_png = pyd.create_png(f="png")
    fp = io.BytesIO(tmp_png)
    img = mpimg.imread(fp, format='png')
    # Show graph
    ax.imshow(img)
    ax.set_title(title + '\n', fontsize=fontsize, loc='left')
    if (show_edge_width) & (ci_type=='directlingam'):
        if graph_type == 'causal':
            ax.annotate(
                'Een dikkere pijl duidt op een sterkere causale relatie', 
                xy=(0,-0.03), xycoords='axes fraction', fontsize=8, fontstyle='italic', ha='left', va='top'
            )
        else:
            ax.annotate(
                'Een dikkere lijn duidt op een grotere correlatie', 
                xy=(0,-0.03), xycoords='axes fraction', fontsize=8, fontstyle='italic', ha='left', va='top'
            )
    if (graph_type == 'causal') & (ci_type=='fci'):
        ax.annotate(
            '-->    Directe invloed (A veroorzaakt B)\no->   Mogelijke invloed met onzekerheid over richting\n<->   Onbekende oorzaak (mogelijk gemeenschappelijke verborgen variabele)\no-o    Geen directe invloed (mogelijk wel verband via verborgen variabele)', 
            xy=(0,-0.03), xycoords='axes fraction', fontsize=8, fontstyle='italic', ha='left', va='top'
        )
    if color_direction:
        if graph_type == 'causal':
            legend_patches = {'Positieve invloed': 'blue', 'Negatieve invloed': 'firebrick'}
            legend_title = 'Causale Richting'
        else:
            legend_patches = {'Positief': 'blue', 'Negatief': 'firebrick'}
            legend_title = 'Correlation'
        ax.legend(handles=[mpatches.Patch(color=legend_patches[label], label=label) for label in legend_patches], title=legend_title, loc='upper left', bbox_to_anchor=(1,1), fontsize=7, title_fontsize=8)    
    return ax


# Function to initialize the graph for background knowledge in PC and FCI
def initialize_graph_for_background_knowledge(data, forbidden_links=None, required_links=None, sample_size=20):
    """
    Initialize graph nodes and define background knowledge for causal discovery algorithms.
    Since defining graph nodes in causal-learn is unintuitive, we run a small subset of  
    the dataset through PC to initialize the graph nodes.

    Parameters:
    - data (pd.DataFrame): Dataset containing the variables.
    - forbidden_links (list of tuples, optional): Forbidden edges as (from_var, to_var). Default is None.
    - required_links (list of tuples, optional): Required edges as (from_var, to_var). Default is None.
    - sample_size (int, optional): Number of rows to sample from the dataset. Default is 20.

    Returns:
    - bk (BackgroundKnowledge): Background knowledge object with constraints applied.
    """
    # Initialize lists if not provided
    if forbidden_links is None:
        forbidden_links = []
    if required_links is None:
        required_links = []    # Run PC on a small subset to initialize nodes
    cg_base = pc(data.sample(min(sample_size, len(data))).values, verbose=0, show_progress=False)
    nodes = cg_base.G.get_nodes()
    
    # Map column names to node indices
    name_to_index = {name: idx for idx, name in enumerate(data.columns)}

    # Initialize BackgroundKnowledge
    bk = BackgroundKnowledge()
    
    # Add forbidden links
    for from_var, to_var in forbidden_links:
        bk.add_forbidden_by_node(nodes[name_to_index[from_var]], nodes[name_to_index[to_var]])
    
    # Add required links
    for from_var, to_var in required_links:
        bk.add_required_by_node(nodes[name_to_index[from_var]], nodes[name_to_index[to_var]])
    return bk
